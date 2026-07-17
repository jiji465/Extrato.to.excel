"""Testes das correções da rodada de melhorias (2026-07-17):

- categorizador: falsos positivos eliminados (das/13/vencimento/nomes curtos);
- valores_brl: parênteses contábeis negativos, sem afetar o "(-)" do BB;
- detector: desempate favorece conta corrente; Caixa sem casar "caixa" genérico;
- excel: fórmula da Conferência protege saldo inicial vazio; SUMIFS limitado;
- rota /converter: upload .pdf.txt (OCR do navegador), erros 400/413 em JSON;
- identificacao: titular, período e nome de arquivo.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest

from extrato.categorizador import classificar, categorizar
from extrato.detector import detectar_banco, detectar_tipo
from extrato.excel import gerar_excel
from extrato.identificacao import (
    extrair_titular, periodo_das_datas, sanitizar_arquivo,
)
from extrato.normalizar import CONTA_CORRENTE, Transacao, valores_brl
from extrato.conversor import ResultadoArquivo


# ---------------------------------------------------------------------------
# Categorizador — falsos positivos corrigidos
# ---------------------------------------------------------------------------
def _cat(descricao: str, favorecido: str = "") -> str:
    t = Transacao(data=None, descricao=descricao, valor=1.0, banco="X",
                  favorecido=favorecido)
    return categorizar([t])[0].categoria


@pytest.mark.parametrize("descricao,favorecido,esperado", [
    # preposição "das" não é a guia DAS
    ("COMPRA DAS PECAS", "", "Compras"),
    # nome de pessoa no favorecido não vira concessionária
    ("PIX RECEBIDO", "TIM SILVA", "Pix recebido"),
    ("", "TIM SILVA", "Não classificado"),
    ("PIX ENVIADO", "MARIA LUZ", "Pix enviado"),
    # "13" em endereço não é 13º salário
    ("TRANSFERENCIA ENVIADA", "RUA 13 DE MAIO COMERCIO", "Transferências"),
    # vencimento de boleto não é folha
    ("PAGAMENTO VENCIMENTO 10/02", "", "Não classificado"),
    # casos que DEVEM continuar classificando
    ("LIQUIDO DE VENCIMENTO", "", "Folha e Salários"),
    ("DEB AUT VIVO", "", "Concessionárias"),
    ("", "SABESP SANEAMENTO", "Concessionárias"),  # nome inequívoco no favorecido
    ("PGDAS SIMPLES NACIONAL", "", "Impostos e Tributos"),
])
def test_categorizador_falsos_positivos(descricao, favorecido, esperado):
    assert _cat(descricao, favorecido) == esperado


def test_classificar_api_publica_preservada():
    assert classificar("PIX ENVIADO Fulano") == "Pix enviado"


# ---------------------------------------------------------------------------
# valores_brl — parênteses contábeis
# ---------------------------------------------------------------------------
def test_valores_brl_parenteses_negativo():
    vals = valores_brl("SALDO (1.234,56)")
    assert len(vals) == 1
    v, neg, _, _ = vals[0]
    assert v == pytest.approx(1234.56) and neg is True


def test_valores_brl_marcador_bb_nao_e_parentese():
    # BB usa "(-)" SOLTO após o número; não envolve => token positivo
    vals = valores_brl("99021 610895 6.000,00 (-)")
    assert len(vals) == 1
    v, neg, _, _ = vals[0]
    assert v == pytest.approx(6000.00) and neg is False


def test_valores_brl_sinais_existentes_preservados():
    assert valores_brl("5,93-")[0][1] is True          # sufixo Santander
    assert valores_brl("-166,40")[0][1] is True        # prefixo Bradesco
    assert valores_brl("- 100,00")[0][1] is False      # hífen solto não é sinal


# ---------------------------------------------------------------------------
# Detector — desempate e assinatura da Caixa
# ---------------------------------------------------------------------------
def test_detector_empate_favorece_conta_corrente():
    # 1 ponto p/ fatura ("fatura") e 1 p/ conta corrente ("extrato"):
    # o desempate deve ir para conta corrente.
    assert detectar_tipo("fatura extrato") == CONTA_CORRENTE


def test_detector_caixa_nao_casa_caixa_generico():
    chave, _ = detectar_banco("saque em caixa eletronico do banco xyz")
    assert chave == "desconhecido"


def test_detector_caixa_por_layout():
    chave, nome = detectar_banco("CAIXA Extrato por período Cliente FULANO")
    assert chave == "caixa"


# ---------------------------------------------------------------------------
# Excel — fórmula protege saldo inicial vazio; SUMIFS com intervalo limitado
# ---------------------------------------------------------------------------
def _res(nome="a.pdf", saldo_inicial=None, saldo_final=100.0, n=1):
    return ResultadoArquivo(
        nome=nome, banco="X", tipo_documento=CONTA_CORRENTE, n_transacoes=n,
        saldo_inicial=saldo_inicial, saldo_final=saldo_final,
    )


def test_excel_conferencia_sem_saldo_inicial_nao_acusa():
    from openpyxl import load_workbook
    trans = [Transacao(data=date(2026, 6, 1), descricao="PIX", valor=100.0,
                       banco="X", arquivo="a.pdf")]
    xlsx = gerar_excel(trans, [_res()])
    wb = load_workbook(BytesIO(xlsx))
    assert set(["Transações", "Conferência", "Resumo", "Categorias",
                "Auditoria"]) <= set(wb.sheetnames)
    formula_i = wb["Conferência"]["I2"].value
    # a fórmula precisa tratar C vazio como "sem referência", não como 0
    assert 'OR($C2=""' in formula_i
    # SUMIFS limitado ao intervalo real, não à coluna inteira
    assert "F:F" not in wb["Conferência"]["D2"].value
    assert "F2:F" in wb["Conferência"]["D2"].value


# ---------------------------------------------------------------------------
# Rota /converter — .pdf.txt (OCR do navegador) e erros em JSON
# ---------------------------------------------------------------------------
_LINHAS_CAIXA = """SALDO ANTERIOR R$ 1.734,59 C
05/06/2026 - 16:35:15 051635 ENVIO TRANSF INTERNET TEV Fulano 1.800,00 D 227,69 C
05/06/2026 - 05:31:20 202605 MENSALIDADE CESTA SERVICO 75,00 D 2.027,69 C
01/06/2026 - 11:16:28 011116 PIX RECEBIDO Credi Shop S A Ins 368,10 C 2.102,69 C
CAIXA Extrato por período
Cliente EMPRESA TESTE LTDA"""


@pytest.fixture()
def cliente():
    from app import app as flask_app
    flask_app.config["TESTING"] = True
    with flask_app.test_client() as c:
        yield c


def test_converter_txt_ocr_navegador(cliente):
    dados = {"arquivos": (BytesIO(_LINHAS_CAIXA.encode("utf-8")),
                          "extrato_caixa.pdf.txt")}
    resp = cliente.post("/converter", data=dados,
                        content_type="multipart/form-data")
    assert resp.status_code == 200
    corpo = resp.get_json()
    rel = corpo["relatorio"][0]
    assert rel["nome"] == "extrato_caixa.pdf"       # .txt some do nome exibido
    assert rel["banco"] == "Caixa Econômica Federal"
    assert rel["transacoes"] == 3
    assert rel["auditoria"]["ok"] is True
    assert corpo["arquivo_b64"]


def test_converter_sem_arquivos_eh_400(cliente):
    resp = cliente.post("/converter", data={}, content_type="multipart/form-data")
    assert resp.status_code == 400
    assert "erro" in resp.get_json()


def test_converter_extensao_invalida_eh_400(cliente):
    dados = {"arquivos": (BytesIO(b"x"), "nota.docx")}
    resp = cliente.post("/converter", data=dados,
                        content_type="multipart/form-data")
    assert resp.status_code == 400


def test_upload_grande_demais_eh_413_json(cliente):
    from app import app as flask_app
    original = flask_app.config["MAX_CONTENT_LENGTH"]
    flask_app.config["MAX_CONTENT_LENGTH"] = 1024   # 1 KB p/ teste
    try:
        dados = {"arquivos": (BytesIO(b"0" * 4096), "grande.pdf")}
        resp = cliente.post("/converter", data=dados,
                            content_type="multipart/form-data")
        assert resp.status_code == 413
        assert "erro" in resp.get_json()
    finally:
        flask_app.config["MAX_CONTENT_LENGTH"] = original


def test_login_next_malicioso_nao_redireciona_fora():
    import app as mod
    if not mod._SENHA:
        mod._SENHA = "teste-senha"                  # habilita o login p/ o teste
    mod.app.config["TESTING"] = True
    try:
        with mod.app.test_client() as c:
            resp = c.post("/login?next=//evil.com", data={"senha": mod._SENHA})
            assert resp.status_code == 302
            assert "evil.com" not in resp.headers["Location"]
    finally:
        mod._SENHA = ""


# ---------------------------------------------------------------------------
# identificacao
# ---------------------------------------------------------------------------
def test_extrair_titular_sicoob():
    texto = "COOP.: 4618-3 / X\nCONTA: 99-0 / EMPRESA EXEMPLO LTDA\nPERÍODO: x"
    assert extrair_titular(texto, "sicoob") == "EMPRESA EXEMPLO LTDA"


def test_extrair_titular_generico_cliente():
    assert extrair_titular("Cliente FULANO DE TAL", "caixa") == "FULANO DE TAL"


def test_periodo_das_datas_mes_unico():
    ds = [date(2026, 6, 1), date(2026, 6, 30)]
    assert periodo_das_datas(ds) == "Junho-2026"


def test_sanitizar_arquivo():
    assert sanitizar_arquivo('Extrato: "X" / Y|Z') == "Extrato X Y Z"
