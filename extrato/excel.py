"""Geração do arquivo Excel a partir das transações normalizadas.

O Excel é AUTOCONFERENTE: além dos dados, traz fórmulas vivas (SUMIFS/IF) que
recalculam sozinhas se o usuário editar um valor. Abas:

- Transações: dados + coluna "Confere" (checa saldo linha a linha) com destaque
  vermelho quando o saldo não bate.
- Conferência: por arquivo, saldo inicial + entradas − saídas = saldo final;
  soma vs. totais informados pelo banco; nº de linhas divergentes. Verde/vermelho.
- Categorias e Resumo: totais por categoria/banco via fórmulas.
- Auditoria: registro da conferência feita na leitura (texto).
"""

from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

from .normalizar import Transacao

_MOEDA = 'R$ #,##0.00;[Red]-R$ #,##0.00'
_DATA_FMT = "DD/MM/YYYY"
_TX = "'Transações'"     # referência à aba de transações nas fórmulas

_HEADER_FILL = PatternFill("solid", fgColor="0F766E")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_VERDE = PatternFill("solid", fgColor="DCFCE7")
_VERDE_FT = Font(bold=True, color="166534")
_VERMELHO = PatternFill("solid", fgColor="FEE2E2")
_VERMELHO_FT = Font(bold=True, color="991B1B")
_LINHA_RUIM = PatternFill("solid", fgColor="FFF1F2")

# Colunas da aba Transações (a ordem define as letras usadas nas fórmulas).
_COLS_TX = ["Data", "Hora", "Descrição", "Favorecido", "Documento", "Valor",
            "Tipo", "Categoria", "Saldo", "Banco", "Arquivo", "Confere"]
# letras: Valor=F, Tipo=G, Categoria=H, Saldo=I, Banco=J, Arquivo=K, Confere=L


def _escrever_cabecalho(ws, colunas: list[str]) -> None:
    for c, titulo in enumerate(colunas, start=1):
        cel = ws.cell(row=1, column=c, value=titulo)
        cel.fill = _HEADER_FILL
        cel.font = _HEADER_FONT
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = _BORDER


# ---------------------------------------------------------------------------
# Aba Transações
# ---------------------------------------------------------------------------
def _aba_transacoes(wb: Workbook, transacoes: list[Transacao]) -> None:
    ws = wb.active
    ws.title = "Transações"
    _escrever_cabecalho(ws, _COLS_TX)

    for t in transacoes:
        ws.append([
            t.data, t.hora, t.descricao, t.favorecido, t.documento,
            t.valor, t.tipo, t.categoria, t.saldo, t.banco, t.arquivo, None,
        ])

    ultima = ws.max_row
    for r in range(2, ultima + 1):
        ws.cell(row=r, column=1).number_format = _DATA_FMT      # Data
        ws.cell(row=r, column=6).number_format = _MOEDA         # Valor
        cel_saldo = ws.cell(row=r, column=9)                    # Saldo
        if cel_saldo.value is not None:
            cel_saldo.number_format = _MOEDA
        # Coluna "Confere": saldo desta linha − saldo da anterior = valor?
        # (só compara dentro do MESMO arquivo e quando há saldo impresso)
        ws.cell(row=r, column=12).value = (
            f'=IF(AND($K{r}=$K{r-1},ISNUMBER($I{r}),ISNUMBER($I{r-1})),'
            f'IF(ABS(($I{r}-$I{r-1})-$F{r})<=0.01,"ok","DIVERGE"),"")'
        )
        for c in range(1, len(_COLS_TX) + 1):
            ws.cell(row=r, column=c).border = _BORDER

    larguras = [12, 10, 34, 30, 16, 14, 10, 20, 14, 18, 26, 11]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if ultima >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLS_TX))}{ultima}"

    if ultima >= 2:
        faixa = f"A2:L{ultima}"
        # linha inteira em vermelho-claro quando o saldo diverge
        ws.conditional_formatting.add(faixa, FormulaRule(
            formula=[f'$L2="DIVERGE"'], fill=_LINHA_RUIM, stopIfTrue=False))
        ws.conditional_formatting.add(f"L2:L{ultima}", CellIsRule(
            operator="equal", formula=['"DIVERGE"'], fill=_VERMELHO, font=_VERMELHO_FT))


# ---------------------------------------------------------------------------
# Aba Conferência (fórmulas vivas + verde/vermelho)
# ---------------------------------------------------------------------------
_COLS_CONF = ["Arquivo", "Banco", "Saldo inicial", "Entradas", "Saídas",
              "Líquido", "Saldo final (calc.)", "Saldo final (extrato)",
              "Confere saldo", "Créditos (extrato)", "Confere créditos",
              "Débitos (extrato)", "Confere débitos", "Saldos linha a linha"]


def _aba_conferencia(wb: Workbook, relatorio: list) -> None:
    ws = wb.create_sheet("Conferência")
    _escrever_cabecalho(ws, _COLS_CONF)

    r = 1
    for res in relatorio:
        if getattr(res, "erro", "") or res.n_transacoes == 0:
            continue
        r += 1
        nome = res.nome
        ws.cell(row=r, column=1, value=nome)                   # A Arquivo
        ws.cell(row=r, column=2, value=res.banco)              # B Banco
        ws.cell(row=r, column=3, value=res.saldo_inicial)      # C Saldo inicial
        # D Entradas / E Saídas por arquivo (SUMIFS pela coluna Arquivo=K)
        ws.cell(row=r, column=4, value=(
            f'=SUMIFS({_TX}!F:F,{_TX}!K:K,$A{r},{_TX}!F:F,">0")'))
        ws.cell(row=r, column=5, value=(
            f'=SUMIFS({_TX}!F:F,{_TX}!K:K,$A{r},{_TX}!F:F,"<0")'))
        ws.cell(row=r, column=6, value=f'=D{r}+E{r}')          # F Líquido
        ws.cell(row=r, column=7, value=f'=C{r}+F{r}')          # G Saldo final calc
        ws.cell(row=r, column=8, value=res.saldo_final)        # H Saldo final extrato
        ws.cell(row=r, column=9, value=(                       # I Confere saldo
            f'=IF(H{r}="","sem referência",'
            f'IF(ABS(G{r}-H{r})<=0.01,"CONFERE","NÃO CONFERE"))'))
        ws.cell(row=r, column=10, value=res.total_creditos)    # J Créditos extrato
        ws.cell(row=r, column=11, value=(                      # K Confere créditos
            f'=IF(J{r}="","—",IF(ABS(D{r}-J{r})<=0.01,"CONFERE","NÃO CONFERE"))'))
        ws.cell(row=r, column=12, value=res.total_debitos)     # L Débitos extrato
        ws.cell(row=r, column=13, value=(                      # M Confere débitos
            f'=IF(L{r}="","—",IF(ABS(-E{r}-L{r})<=0.01,"CONFERE","NÃO CONFERE"))'))
        ws.cell(row=r, column=14, value=(                      # N Saldos linha a linha
            f'=IF(COUNTIFS({_TX}!K:K,$A{r},{_TX}!L:L,"DIVERGE")=0,"CONFERE",'
            f'"VER "&COUNTIFS({_TX}!K:K,$A{r},{_TX}!L:L,"DIVERGE")&" linha(s)")'))

        for col in (3, 4, 5, 6, 7, 8, 10, 12):
            ws.cell(row=r, column=col).number_format = _MOEDA
        for col in range(1, len(_COLS_CONF) + 1):
            ws.cell(row=r, column=col).border = _BORDER

    ult = r
    larguras = [26, 16, 14, 14, 14, 14, 16, 16, 15, 16, 15, 16, 15, 18]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # verde/vermelho nas colunas de veredito (I, K, M, N)
    if ult >= 2:
        for col in ("I", "K", "M", "N"):
            faixa = f"{col}2:{col}{ult}"
            ws.conditional_formatting.add(faixa, CellIsRule(
                operator="equal", formula=['"CONFERE"'], fill=_VERDE, font=_VERDE_FT))
            ws.conditional_formatting.add(faixa, CellIsRule(
                operator="equal", formula=['"NÃO CONFERE"'], fill=_VERMELHO, font=_VERMELHO_FT))
            ws.conditional_formatting.add(faixa, FormulaRule(
                formula=[f'ISNUMBER(SEARCH("VER",{col}2))'], fill=_VERMELHO, font=_VERMELHO_FT))


# ---------------------------------------------------------------------------
# Aba Categorias (fórmulas)
# ---------------------------------------------------------------------------
def _aba_categorias(wb: Workbook, transacoes: list[Transacao]) -> None:
    ws = wb.create_sheet("Categorias")
    _escrever_cabecalho(ws, ["Categoria", "Entradas", "Saídas", "Total", "Nº lançamentos"])

    cats = sorted({(t.categoria or "Não classificado") for t in transacoes})
    for i, cat in enumerate(cats):
        r = i + 2
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=f'=SUMIFS({_TX}!F:F,{_TX}!H:H,$A{r},{_TX}!F:F,">0")')
        ws.cell(row=r, column=3, value=f'=SUMIFS({_TX}!F:F,{_TX}!H:H,$A{r},{_TX}!F:F,"<0")')
        ws.cell(row=r, column=4, value=f'=B{r}+C{r}')
        ws.cell(row=r, column=5, value=f'=COUNTIFS({_TX}!H:H,$A{r})')
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).number_format = _MOEDA
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = _BORDER

    for i, w in enumerate([26, 16, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Aba Resumo por banco (fórmulas)
# ---------------------------------------------------------------------------
def _aba_resumo(wb: Workbook, transacoes: list[Transacao]) -> None:
    ws = wb.create_sheet("Resumo")
    _escrever_cabecalho(ws, ["Banco", "Entradas", "Saídas", "Saldo do período", "Nº lançamentos"])

    bancos = sorted({t.banco for t in transacoes})
    for i, banco in enumerate(bancos):
        r = i + 2
        ws.cell(row=r, column=1, value=banco)
        ws.cell(row=r, column=2, value=f'=SUMIFS({_TX}!F:F,{_TX}!J:J,$A{r},{_TX}!F:F,">0")')
        ws.cell(row=r, column=3, value=f'=SUMIFS({_TX}!F:F,{_TX}!J:J,$A{r},{_TX}!F:F,"<0")')
        ws.cell(row=r, column=4, value=f'=B{r}+C{r}')
        ws.cell(row=r, column=5, value=f'=COUNTIFS({_TX}!J:J,$A{r})')
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).number_format = _MOEDA
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = _BORDER

    for i, w in enumerate([22, 16, 16, 18, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Aba Auditoria (registro textual da conferência feita na leitura)
# ---------------------------------------------------------------------------
def _aba_auditoria(wb: Workbook, relatorio: list) -> None:
    ws = wb.create_sheet("Auditoria")
    _escrever_cabecalho(ws, ["Arquivo", "Banco", "Tipo", "Transações",
                             "Conferência", "Detalhes"])

    for r in relatorio:
        aud = getattr(r, "auditoria", None)
        tipo = getattr(r, "tipo_label", "—")
        if r.erro:
            status, detalhes = "ERRO", r.erro
        elif aud is not None:
            status = aud.resumo
            detalhes = " | ".join(aud.mensagens)
        else:
            status, detalhes = "—", ""
        linha = ws.max_row + 1
        ws.append([r.nome, r.banco, tipo, r.n_transacoes, status, detalhes])
        cel = ws.cell(row=linha, column=5)
        if aud is not None and aud.conferido and aud.ok and not r.erro:
            cel.font = Font(bold=True, color="16A34A")
        elif r.erro or (aud is not None and aud.conferido and not aud.ok):
            cel.font = Font(bold=True, color="DC2626")
        else:
            cel.font = Font(bold=True, color="B45309")

    for row in ws.iter_rows(min_row=2):
        for cel in row:
            cel.border = _BORDER
            cel.alignment = Alignment(vertical="top", wrap_text=(cel.column == 6))
    for i, w in enumerate([32, 18, 16, 12, 22, 66], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def gerar_excel(transacoes: Iterable[Transacao], relatorio: list | None = None) -> bytes:
    """Monta o workbook (com fórmulas de conferência) e devolve os bytes do .xlsx."""
    transacoes = list(transacoes)
    wb = Workbook()
    _aba_transacoes(wb, transacoes)
    if relatorio:
        _aba_conferencia(wb, relatorio)
    _aba_resumo(wb, transacoes)
    _aba_categorias(wb, transacoes)
    if relatorio:
        _aba_auditoria(wb, relatorio)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
